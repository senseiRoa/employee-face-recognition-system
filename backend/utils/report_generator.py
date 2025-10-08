"""
Utilidad para generar reportes en diferentes formatos
"""

import io
import pandas as pd
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Tuple
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib.backends.backend_agg import FigureCanvasAgg
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference, LineChart

from models import Employee, AccessLog, Warehouse


class ReportGenerator:
    """Generador de reportes en múltiples formatos"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Configurar estilos personalizados para PDF"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomSubtitle',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=20,
            textColor=colors.darkgreen
        ))
    
    def generate_attendance_report_pdf(
        self, 
        data: List[Dict[str, Any]], 
        title: str,
        period: str
    ) -> bytes:
        """Generar reporte de asistencia en PDF"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Título
        title_para = Paragraph(f"{title}", self.styles['CustomTitle'])
        elements.append(title_para)
        elements.append(Spacer(1, 12))
        
        # Período
        period_para = Paragraph(f"Período: {period}", self.styles['CustomSubtitle'])
        elements.append(period_para)
        elements.append(Spacer(1, 12))
        
        # Fecha de generación
        generated_para = Paragraph(
            f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            self.styles['Normal']
        )
        elements.append(generated_para)
        elements.append(Spacer(1, 20))
        
        if data:
            # Tabla de datos
            table_data = [['Empleado', 'Almacén', 'Entradas', 'Último Acceso']]
            
            for row in data:
                table_data.append([
                    row.get('employee_name', 'N/A'),
                    row.get('warehouse_name', 'N/A'),
                    str(row.get('total_entries', 0)),
                    row.get('last_access', 'N/A')
                ])
            
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
        else:
            no_data_para = Paragraph("No se encontraron datos para el período seleccionado.", self.styles['Normal'])
            elements.append(no_data_para)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_attendance_report_excel(
        self, 
        data: List[Dict[str, Any]], 
        title: str,
        period: str
    ) -> bytes:
        """Generar reporte de asistencia en Excel"""
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Asistencia"
        
        # Configurar estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center")
        
        # Título
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Período
        ws['A2'] = f"Período: {period}"
        ws['A2'].font = Font(bold=True)
        ws.merge_cells('A2:D2')
        
        # Fecha de generación
        ws['A3'] = f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A3:D3')
        
        # Headers
        headers = ['Empleado', 'Almacén', 'Entradas', 'Último Acceso']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Datos
        if data:
            for row_idx, row_data in enumerate(data, 6):
                ws.cell(row=row_idx, column=1, value=row_data.get('employee_name', 'N/A'))
                ws.cell(row=row_idx, column=2, value=row_data.get('warehouse_name', 'N/A'))
                ws.cell(row=row_idx, column=3, value=row_data.get('total_entries', 0))
                ws.cell(row=row_idx, column=4, value=row_data.get('last_access', 'N/A'))
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_attendance_report_csv(
        self, 
        data: List[Dict[str, Any]], 
        title: str,
        period: str
    ) -> bytes:
        """Generar reporte de asistencia en CSV"""
        # Crear DataFrame
        if data:
            df = pd.DataFrame(data)
            # Renombrar columnas para mejor legibilidad
            df = df.rename(columns={
                'employee_name': 'Empleado',
                'warehouse_name': 'Almacén',
                'total_entries': 'Entradas',
                'last_access': 'Último Acceso'
            })
        else:
            df = pd.DataFrame(columns=['Empleado', 'Almacén', 'Entradas', 'Último Acceso'])
        
        # Convertir a CSV
        buffer = io.StringIO()
        buffer.write(f"# {title}\\n")
        buffer.write(f"# Período: {period}\\n")
        buffer.write(f"# Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M')}\\n")
        buffer.write("\\n")
        
        df.to_csv(buffer, index=False, encoding='utf-8')
        
        # Convertir a bytes
        csv_content = buffer.getvalue()
        return csv_content.encode('utf-8')
    
    def generate_chart_data(
        self, 
        data: List[Dict[str, Any]], 
        chart_type: str = 'line'
    ) -> str:
        """Generar gráfico y retornar como base64"""
        if not data:
            return ""
        
        # Configurar matplotlib para no mostrar ventanas
        plt.switch_backend('Agg')
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Extraer datos para el gráfico
        if chart_type == 'line':
            dates = [item['date'] for item in data]
            values = [item['count'] for item in data]
            
            ax.plot(dates, values, marker='o', linewidth=2, markersize=6)
            ax.set_xlabel('Fecha')
            ax.set_ylabel('Cantidad')
            ax.set_title('Tendencia de Asistencia')
            
            # Formatear fechas en el eje X
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m'))
            ax.xaxis.set_major_locator(mdates.DayLocator(interval=7))
            plt.xticks(rotation=45)
            
        elif chart_type == 'bar':
            labels = [item['label'] for item in data]
            values = [item['value'] for item in data]
            
            ax.bar(labels, values, color='skyblue', alpha=0.7)
            ax.set_xlabel('Categoría')
            ax.set_ylabel('Cantidad')
            ax.set_title('Distribución por Categoría')
            
            plt.xticks(rotation=45)
        
        plt.tight_layout()
        
        # Convertir a base64
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
        buffer.seek(0)
        
        image_base64 = base64.b64encode(buffer.getvalue()).decode()
        plt.close(fig)
        
        return image_base64
    
    def get_report_filename(self, report_type: str, format_type: str, timestamp: datetime = None) -> str:
        """Generar nombre de archivo para el reporte"""
        if timestamp is None:
            timestamp = datetime.now()
        
        date_str = timestamp.strftime('%Y%m%d_%H%M%S')
        
        format_extension = {
            'pdf': 'pdf',
            'excel': 'xlsx',
            'csv': 'csv'
        }
        
        ext = format_extension.get(format_type, 'pdf')
        return f"{report_type}_{date_str}.{ext}"
    
    def get_content_type(self, format_type: str) -> str:
        """Obtener content type para el formato especificado"""
        content_types = {
            'pdf': 'application/pdf',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'csv': 'text/csv'
        }
        return content_types.get(format_type, 'application/octet-stream')